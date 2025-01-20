import pandas as pd
from playwright.async_api import async_playwright
import asyncio
import nest_asyncio
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

# Configurations
input_file = 'Teste.xlsx'  # Input Excel file containing RUCs
url = 'https://portalunico.siscomex.gov.br/due/x/#/consulta/consulta-filtro?perfil=publico'

# Read the input file and limit to the first 70 rows
df = pd.read_excel(input_file).iloc[0:71]


def apply_styles_to_entire_sheet(file_path):
    # Load the Excel workbook
    wb = load_workbook(file_path)
    ws = wb.active  # Select the active worksheet

    # Define styles
    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    pink_fill = PatternFill(start_color="e75480", end_color="e75480", fill_type="solid")
    green_fill = PatternFill(start_color="90ee90", end_color="90ee90", fill_type="solid")
    white_font = Font(color="FFFFFF")
    lighter_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Iterate through all rows and columns
    for row in range(1, ws.max_row + 1):  # Rows from 1 to max_row
        for col in range(1, ws.max_column + 1):  # Columns from 1 to max_column
            cell = ws.cell(row=row, column=col)

            # Apply styles for the header row
            if row == 1:  
                if col == 1:  # First column header (e.g., "RUC")
                    cell.fill = pink_fill
                else:  # Other headers
                    cell.fill = black_fill
                cell.font = white_font
            elif col == 1:  # Apply light green for any "RUC" column
                if row % 2 == 0 and row < 71:
                    cell.fill = green_fill
            # Apply borders to all cells
            cell.border = lighter_border

    wb.save(file_path)


def process_excel(df):
    result = pd.DataFrame()

    # Iterate through unique RUC values
    for ruc in df['RUC'].unique():
        # Filter the rows for the current RUC
        ruc_data = df[df['RUC'] == ruc]
        
        # Create a list to hold alternating 'data/hora' and 'evento' values
        alternating_columns = []
        for date_time, event in zip(ruc_data['Data / hora'], ruc_data['Evento']):
            alternating_columns.append(('Data / Hora', date_time))
            alternating_columns.append(('Evento', event))
        
        # Convert alternating list into a dictionary to create one row
        row = {'RUC': ruc}
        for i, (key, value) in enumerate(alternating_columns):
            row[f'{key}_{i // 2 + 1}'] = value
        
        # Append the row to the result DataFrame
        result = pd.concat([result, pd.DataFrame([row])], ignore_index=True)
    
    return result


def update_ruc_in_place(original_file, result):
    # Read the original file with only the first column (RUC)
    original_df = pd.read_excel(original_file, usecols=[0])  # Load only the RUC column
    original_df = original_df.drop_duplicates().reset_index(drop=True)  # Remove duplicates if any

    # Ensure the result has the same RUCs in the original order
    result_dict = result.set_index('RUC').to_dict(orient='index')

    # Update rows in the original DataFrame with processed data
    for idx, ruc in original_df['RUC'].items():  # Replace iteritems() with items()
        if ruc in result_dict:
            for col, value in result_dict[ruc].items():
                original_df.loc[idx, col] = value

    return original_df


async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)  # Set headless=False to see the browser
        page = await browser.new_page()

        # Navigate to the URL
        await page.goto(url, timeout=120000)  # Increase to 2 minutes

        # Initialize a list to store extracted data
        all_extracted_data = []

        # Function to check for CAPTCHA and handle it
        async def handle_captcha(page):
            try:
                captcha_iframe = page.locator('iframe[title="Widget containing checkbox for hCaptcha security challenge"]')
                if await captcha_iframe.is_visible():
                    print("hCaptcha detected. Please solve the CAPTCHA manually in the browser.")
                    input("Solve the CAPTCHA manually and press Enter to continue...")
                else:
                    print("No CAPTCHA detected.")
            except Exception as e:
                print(f"Error handling CAPTCHA: {e}")

        # Loop through each RUC in the input DataFrame
        for _ , row in df.iterrows():
            ruc = row['RUC']
            print(f"Processing RUC: {ruc}")

            try:
                # Select the RUC option
                await page.locator('#opt2 label').click()

                # Enter the RUC in the input field
                await page.locator('#nrRuc').fill(ruc)

                # Handle CAPTCHA if present
                await handle_captcha(page)

                # Click the "Consultar" button
                await page.get_by_role("button", name="Consultar").click()
                print("Consulted successfully.")

                # Wait for the results table to load
                await page.wait_for_selector("table", timeout=60000)

                # Locate and click the DU-E link dynamically
                du_e_link_locator = page.locator("table a")
                du_e_links = await du_e_link_locator.all_text_contents()

                if du_e_links:
                    print(f"DU-E links found: {du_e_links}")
                    await du_e_link_locator.first.click()
                    await page.wait_for_selector("table", timeout=120000)
                    print("Clicked on DU-E link.")

                    # Extract data from the new table
                    rows = await page.locator('table tbody tr').all()
                    data = []
                    for row in rows:
                        cells = await row.locator('td').all_text_contents()
                        data.append({
                            "RUC": ruc,
                            "Data / hora": cells[0] if len(cells) > 0 else None,
                            "Evento": cells[1] if len(cells) > 1 else None,
                        })

                    all_extracted_data.extend(data)
                else:
                    print("No DU-E links found for this RUC.")
                    
                await asyncio.sleep(3)

                # Return to the main search page
                await page.go_back()
                await asyncio.sleep(2)
            
            except Exception as e:
                print(f"An error occurred while processing RUC {ruc}: {e}")

        
        processed_result = process_excel(all_extracted_data)
        final_df = update_ruc_in_place('Teste.xlsx', processed_result)
        
        clean_columns = {}
        for col in final_df.columns:
            if 'Data / Hora_' in col:
                clean_columns[col] = 'Data / Hora'
            elif 'Evento_' in col:
                clean_columns[col] = 'Evento'
            else:
                clean_columns[col] = col
        final_df = final_df.rename(columns=clean_columns)

        # Save the updated file
        final_df.to_excel("Teste.xlsx", index=False)
        
        apply_styles_to_entire_sheet('Teste.xlsx')
                
        # Close the browser
        await browser.close()
        
        
# Apply the nested event loop fix
nest_asyncio.apply()

if __name__ == "__main__":
    asyncio.run(main())
